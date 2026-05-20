"""
Generates docs/QA-TestPlan.xlsx — test scenarios for the tester, grouped by feature.

Sheets:
  - Summary          : feature roll-up with counts by priority/status
  - Test Cases       : all scenarios, one row per scenario (filterable)
  - Status Legend    : meanings for Status / Priority columns

Run:
  python scripts/generate_test_plan.py
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# ----- styling -----
HDR_FILL   = PatternFill("solid", fgColor="1F4E78")
HDR_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT  = Font(name="Calibri", size=10)
WRAP       = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIO_FILLS = {
    "High":   PatternFill("solid", fgColor="F8CBAD"),
    "Medium": PatternFill("solid", fgColor="FFE699"),
    "Low":    PatternFill("solid", fgColor="C6E0B4"),
}

# ----- test cases -----
# (Feature, Sub-Feature, Scenario, Steps, Expected, Priority)
CASES = [
    # ============ AUTH & REGISTRATION ============
    ("Auth", "Send OTP", "Send OTP to a valid Saudi mobile number for registration",
     "POST /api/auth/send-otp with mobile=+9665XXXXXXXX",
     "200 OK with pinId; OTP delivered via WhatsApp", "High"),
    ("Auth", "Send OTP", "Send OTP to an already-registered mobile",
     "POST /api/auth/send-otp using a mobile that is already approved",
     "400 with Auth.MobileAlreadyRegistered (or equivalent error)", "High"),
    ("Auth", "Send OTP", "Send OTP to invalid mobile format",
     "POST /api/auth/send-otp with mobile=12345",
     "400 ProblemDetails — mobile validation error", "Medium"),
    ("Auth", "Send OTP", "Rate-limit OTP requests for same mobile",
     "Send OTP 5+ times in a row to same mobile within rate-limit window",
     "Subsequent calls return 429 / rate-limited error", "Medium"),

    ("Auth", "Register Shop Owner", "Register Shop Owner (multipart) with valid OTP + ShopImage",
     "POST /api/auth/register/shop-owner [FromForm] with PinId, Otp, all required fields, ShopImage file",
     "201 with user info; user RegistrationStatus=PendingSalesman; ShopData created (overwrites Seller's if exists)", "High"),
    ("Auth", "Register Shop Owner", "Register with CustomerCode that already has an approved ShopOwner",
     "Submit registration using CustomerCode already linked to an approved ShopOwner",
     "400 — only one ShopOwner allowed per CustomerCode", "High"),
    ("Auth", "Register Shop Owner", "Register with invalid VAT (not 15 digits, doesn't start/end with 3)",
     "POST /api/auth/register/shop-owner with VAT=123",
     "400 ProblemDetails listing VAT validation error", "Medium"),
    ("Auth", "Register Shop Owner", "Register with invalid CRN (not 10 digits)",
     "POST /api/auth/register/shop-owner with CRN=ABC",
     "400 ProblemDetails listing CRN validation error", "Medium"),
    ("Auth", "Register Shop Owner", "Register with name containing numbers/symbols",
     "POST /api/auth/register/shop-owner with FirstName='Ahmed123'",
     "400 — Arabic/Latin letters only", "Medium"),
    ("Auth", "Register Shop Owner", "Register with wrong OTP code",
     "POST register with Otp=000000",
     "400 — invalid OTP; previous validation errors must be reported BEFORE OTP check", "High"),

    ("Auth", "Register Seller", "Register Seller for an existing CustomerCode (data already provided by ShopOwner)",
     "POST /api/auth/register/seller [FromForm] for CustomerCode whose ShopData exists",
     "201 — Seller created without overwriting ShopData", "High"),
    ("Auth", "Register Seller", "First Seller for a CustomerCode (no ShopOwner yet)",
     "POST /api/auth/register/seller for unused CustomerCode with all shop fields + image",
     "201 — Seller creates ShopData (first-come-first-served)", "High"),
    ("Auth", "Register Seller", "Subsequent Seller for same CustomerCode skips shop fields",
     "POST /api/auth/register/seller after another Seller already provided shop data",
     "201 — shop data unchanged", "Medium"),

    ("Auth", "Register Technician", "Register Technician with NationalAddress incl. District",
     "POST /api/auth/register/technician [FromBody] with all required fields",
     "201 — user created, RegistrationStatus=PendingSalesman", "High"),

    ("Auth", "Login", "Login with approved user credentials (mobile + OTP)",
     "POST /api/auth/send-otp then /api/auth/login with PinId + Otp",
     "200 with accessToken + refreshToken", "High"),
    ("Auth", "Login", "Login as a Rejected user",
     "Login attempt by a user with RegistrationStatus=Rejected",
     "403 UserRejected", "High"),
    ("Auth", "Login", "Login as a still-Pending user",
     "Login attempt while user is PendingSalesman or PendingZoneManager",
     "403 UserNotApproved", "High"),
    ("Auth", "Login", "Login as a disabled user",
     "Admin disables the user, user attempts login",
     "403 — login blocked", "High"),
    ("Auth", "Login", "Mock OTP '123456' should NOT work in Staging",
     "Staging environment, login with Otp=123456",
     "Twilio Verify rejects — login fails (only Dev/UAT bypass mock)", "High"),

    ("Auth", "Refresh Token", "Refresh with valid token returns new pair",
     "POST /api/auth/refresh with current refreshToken",
     "200 with new accessToken + new refreshToken; old refresh token revoked", "High"),
    ("Auth", "Refresh Token", "Refresh after user is Rejected",
     "User is rejected; mobile uses old refresh token",
     "403 UserRejected (no new tokens issued)", "High"),
    ("Auth", "Refresh Token", "Refresh after user is Disabled",
     "Admin disables user; old refresh token used",
     "403 / refresh blocked — disable revokes all refresh tokens", "High"),
    ("Auth", "Refresh Token", "Reuse a revoked refresh token (theft detection)",
     "Use a refresh token that was already used once",
     "403 / token reuse detected; user's whole token family revoked", "High"),

    ("Auth", "Admin Login", "Admin login with username/password",
     "POST /api/admin/auth/login with username=admin password=Raed@2026",
     "200 with admin tokens", "High"),
    ("Auth", "Admin Login", "Non-admin user attempts admin login",
     "POST /api/admin/auth/login with a Seller's credentials",
     "401/403 — endpoint restricted to SystemAdmin role", "High"),

    # ============ APPROVAL FLOW ============
    ("Approval", "SM Approve", "SM approves user in their assigned city",
     "Login as SM, POST /api/approval/{userId}/approve for a PendingSalesman user in SM's city",
     "200 — user moves to PendingZoneManager", "High"),
    ("Approval", "SM Approve", "SM tries to approve user from another SM's city",
     "Different SM attempts approval",
     "403 NotAuthorizedToApprove", "High"),
    ("Approval", "SM Approve", "Dual-role SM+ZM approves — skips ZM step",
     "Approver has SM role for the city AND ZM role for the region",
     "User jumps directly to Approved (one step)", "High"),
    ("Approval", "ZM Approve", "ZM approves PendingZoneManager user",
     "Login as ZM for the region, approve user",
     "200 — RegistrationStatus=Approved; reward triggers if invited", "High"),
    ("Approval", "ZM Approve", "ZM tries to approve user outside their region",
     "ZM of region A approves user in region B",
     "403 NotAuthorizedToApprove", "High"),
    ("Approval", "ZM Approve", "Region with no ZoneManager assigned",
     "Try to approve in a region whose Region.ZoneManagerId is null",
     "400 NoZoneManagerForRegion (this should be unreachable post-S19 invariant)", "Medium"),
    ("Approval", "Reject", "SM rejects a registration",
     "POST /api/approval/{userId}/reject with reason",
     "200 — RegistrationStatus=Rejected; user's refresh tokens revoked in same tx", "High"),
    ("Approval", "List/Search (SM)", "SM list pending approvals",
     "GET /api/approval/list?status=Pending as SM",
     "Only users in SM's cities, status=PendingSalesman", "High"),
    ("Approval", "List/Search (ZM)", "ZM list pending approvals",
     "GET /api/approval/list?status=Pending as ZM",
     "Only users in ZM's region, status=PendingZoneManager", "High"),
    ("Approval", "List/Search", "Filter by tab All/Approved/Rejected",
     "GET /api/approval/list?status=Approved (and Rejected, All) with search query",
     "Correct filtering + search by name/mobile", "Medium"),

    # ============ PROFILE ============
    ("Profile", "Get", "Get my profile (Seller/Technician/ShopOwner)",
     "GET /api/profile with valid token",
     "200 with profile data + Points (wallet balance)", "High"),
    ("Profile", "Get", "Get my profile as SalesMan / ZoneManager",
     "GET /api/profile as SM or ZM",
     "200 with profile data + Points=null (no wallet)", "High"),
    ("Profile", "Update Photo", "Upload new profile photo",
     "PUT /api/profile/photo with valid image file",
     "200; new photo URL returned + persisted", "Medium"),
    ("Profile", "Update Photo", "Upload non-image file as photo",
     "PUT /api/profile/photo with a .pdf",
     "400 — invalid file type", "Low"),
    ("Profile", "Delete Account", "Self-delete account (Seller/Technician/ShopOwner)",
     "DELETE /api/profile",
     "200 — IsAccountDeleted=true; all refresh tokens revoked; FcmToken cleared", "High"),
    ("Profile", "Delete Account", "Self-delete blocked for SalesMan/ZoneManager",
     "DELETE /api/profile as SM or ZM",
     "403/400 — staff cannot self-delete", "Medium"),

    # ============ SCANNING & WALLET ============
    ("Scan", "Seller Scan", "Seller scans a fresh barcode",
     "POST /api/scan with barcode of an Available product",
     "200 — points credited to seller's wallet; barcode → SellerScanned", "High"),
    ("Scan", "Technician Scan", "Technician scans a SellerScanned barcode",
     "POST /api/scan as Technician on a SellerScanned barcode",
     "200 — tech earns points + seller earns deferred points; barcode → Consumed", "High"),
    ("Scan", "Duplicate Scan", "Same scanner scans a barcode they already scanned",
     "POST /api/scan twice in same role",
     "400 — composite unique (BarcodeId + ScannerRole) violation", "High"),
    ("Scan", "Invalid Barcode", "Scan a non-existent barcode",
     "POST /api/scan with random NanoID",
     "404 — barcode not found", "Medium"),
    ("Scan", "Invalid Barcode", "Scan a Consumed barcode",
     "POST /api/scan on already-consumed barcode",
     "400 — barcode no longer redeemable", "High"),
    ("Scan", "Concurrency", "Two scanners hit same barcode simultaneously",
     "Concurrent POST /api/scan from two devices on same barcode",
     "One succeeds; other gets ConcurrencyConflict 409 (RowVersion)", "Medium"),
    ("Scan", "Wallet First Scan", "User's first ever scan (wallet doesn't exist)",
     "POST /api/scan as a user with no wallet row",
     "200 — wallet pre-created outside tx; scan succeeds", "High"),
    ("Scan", "ShopOwner Scan", "ShopOwner attempts to scan",
     "POST /api/scan as ShopOwner",
     "403 UnauthorizedRole — ShopOwners don't scan", "High"),
    ("Scan", "History", "Scan history paginated",
     "GET /api/scan/history?page=1&pageSize=20",
     "200 with paginated results", "Medium"),
    ("Wallet", "Balance", "Get wallet balance",
     "GET /api/wallet/balance",
     "200 with Balance (points, decimal 10,1) + SarBalance (decimal 10,2)", "High"),
    ("Wallet", "Transactions", "List wallet transactions paginated",
     "GET /api/wallet/transactions?page=1&pageSize=20",
     "200 paginated; each item has Type, Points, SarAmount, SarRate, ReferenceId", "High"),

    # ============ REDEMPTION ============
    ("Redemption", "Cash Request", "Submit cash redemption request when balance >= MinimumRedemptionPoints",
     "POST /api/redemption/cash with valid points",
     "201; cash OTP sent via WhatsApp at request creation; transactional rollback if WhatsApp fails", "High"),
    ("Redemption", "Cash Request", "Submit cash request below MinimumRedemptionPoints",
     "POST /api/redemption/cash with points < setting",
     "400 — minimum points not met (check reads RewardSettings, not hardcoded)", "High"),
    ("Redemption", "Cash Request", "Submit cash request with 0 / negative points",
     "POST /api/redemption/cash with points=0",
     "400 from validator", "Medium"),
    ("Redemption", "Cash Request", "Cash OTP delivery failure rolls back the request",
     "Force WhatsApp send to fail (UAT real Twilio with bad template)",
     "502 — request not persisted; user retries cleanly", "High"),
    ("Redemption", "Resend Cash OTP", "Resend cash OTP after 30s cooldown",
     "POST /api/redemption/resend-cash-otp after waiting 30s+",
     "200 — new OTP sent; attempts/expiry reset", "High"),
    ("Redemption", "Resend Cash OTP", "Resend within 30s cooldown",
     "POST /api/redemption/resend-cash-otp again immediately",
     "429 — cooldown active", "Medium"),
    ("Redemption", "Resend Cash OTP", "Resend for non-existent / closed request",
     "POST resend on a confirmed or non-existent request",
     "404", "Medium"),
    ("Redemption", "Bank Transfer", "Submit bank transfer request",
     "POST /api/redemption/bank-transfer with IBAN + amount",
     "201 — request created (no OTP path)", "High"),
    ("Redemption", "Confirm Cash Handover", "SM confirms handover with valid OTP",
     "SM enters cash OTP from user",
     "200 — request status=Completed; points debited; SAR snapshot stored", "High"),
    ("Redemption", "Confirm Cash Handover", "SM tries to confirm in another region (IDOR)",
     "Different SM uses guessed RedemptionRequestId",
     "403 — city/region scope check denies it", "High"),
    ("Redemption", "Confirm Cash Handover", "Confirm with wrong OTP",
     "SM enters incorrect OTP",
     "400 — invalid OTP; attempt counter increments", "High"),
    ("Redemption", "Confirm Cash Handover", "Confirm with expired OTP (>14 days)",
     "Wait past OTP expiry then attempt confirm",
     "400 — OTP expired", "Medium"),
    ("Redemption", "Approval Levels", "ZM approval step",
     "Cash request → SM approval → ZM approval → Admin approval",
     "Each level transitions correctly; only the right approver can act", "High"),
    ("Redemption", "Pending Queue (Dual-role)", "SM+ZM dual-role sees combined pending",
     "Login as user with both roles, fetch pending list",
     "Combined SM-pending + ZM-pending across scoped cities/region", "Medium"),

    # ============ INVITATIONS ============
    ("Invitation", "Get Code/QR", "Fetch my invitation code + QR",
     "GET /api/invitation",
     "200 with InvitationCode + QR PNG/base64", "High"),
    ("Invitation", "Use Code", "Register using a valid invitation code",
     "POST /api/auth/register/seller with InvitationCode of an existing user",
     "201; invitee linked via InvitedByUserId", "High"),
    ("Invitation", "Reward Crediting", "Invitee gets reward at signup approval",
     "Invited user gets approved → check invitee wallet",
     "InviteeRewardPoints credited (default 50)", "High"),
    ("Invitation", "Reward Crediting", "Inviter gets reward when invitee approved",
     "Invitee approval triggers reward",
     "InviterRewardPoints credited to inviter (default 100)", "High"),
    ("Invitation", "Reward Cap", "Inviter cap stops at 20 rewarded invitations",
     "Approve invitee #21 for the same inviter",
     "Inviter gets NO reward; InviterRewardCount stays at 20; invitee still gets their reward", "High"),
    ("Invitation", "Reward Cap (Concurrency)", "Two concurrent approvals for same inviter near cap",
     "Approve invitee #20 and #21 in parallel for same inviter",
     "Atomic ExecuteUpdateAsync ensures exactly one of them gets credited; counter never exceeds 20", "High"),
    ("Invitation", "Inviter Disabled", "Invitee approved while inviter is disabled",
     "Disable inviter, then approve their invitee",
     "Invitee still gets reward; inviter does NOT", "High"),
    ("Invitation", "Inviter Rejected", "Invitee approved while inviter is Rejected",
     "Set inviter to Rejected, approve invitee",
     "Invitee gets reward; inviter does not", "Medium"),
    ("Invitation", "Bad Code", "Register with invalid invitation code",
     "Use a non-existent code at registration",
     "400 — invalid invitation code", "Medium"),

    # ============ DASHBOARD ============
    ("Dashboard", "Mobile Home", "Get user dashboard (Seller/Technician)",
     "GET /api/dashboard",
     "200 with user info + wallet + reward settings + recent 10 transactions", "High"),
    ("Dashboard", "Mobile Home", "Get dashboard as ShopOwner",
     "GET /api/dashboard as ShopOwner",
     "200 — no TotalScans field; ShopOwner-specific shape", "Medium"),

    # ============ NOTIFICATIONS ============
    ("Notifications", "List", "List my notifications paginated",
     "GET /api/notifications?page=1",
     "200 paginated", "High"),
    ("Notifications", "Unread Count", "Get unread count",
     "GET /api/notifications/unread-count",
     "200 with integer count", "Medium"),
    ("Notifications", "Mark Read", "Mark single notification as read",
     "POST /api/notifications/{id}/read",
     "200; unread count decremented", "Medium"),
    ("Notifications", "Mark All Read", "Mark all my notifications as read",
     "POST /api/notifications/read-all",
     "200; all my notifications IsRead=true", "Medium"),
    ("Notifications", "Delete", "Soft-delete a notification",
     "DELETE /api/notifications/{id}",
     "200; not in subsequent list", "Low"),
    ("Notifications", "Preferences", "Update notification preferences",
     "PUT /api/notifications/preferences with toggles",
     "200; preferences persisted", "Medium"),
    ("Notifications", "FCM Register", "Register FCM device token",
     "POST /api/notifications/register-device with token",
     "200; FcmToken stored on user", "High"),
    ("Notifications", "FCM Unregister", "Unregister FCM device",
     "POST /api/notifications/unregister-device",
     "200; FcmToken cleared", "Medium"),
    ("Notifications", "FCM On Disable", "Disabled user no longer receives push",
     "Admin disables user → trigger any push notification path",
     "User's FcmToken cleared on disable; no push delivered", "High"),
    ("Notifications", "Broadcast Excludes Deleted", "Admin sends to all — deleted users excluded",
     "Admin broadcasts; check delivery list",
     "Users with IsAccountDeleted=true skipped; only active users targeted", "High"),

    # ============ TRADER MAP / SHOPS ============
    ("Trader Map", "List", "List shops on map (with registered ShopData only)",
     "GET /api/shops/map",
     "200 with shops; each item has shortAddress, street, district, buildingNumber, phone (registering user mobile)", "High"),
    ("Trader Map", "Filter", "Map filtered to only customers with registered ShopData",
     "GET /api/shops/map — confirm orphan ErpCustomers excluded",
     "Items only include shops where ShopData exists", "Medium"),

    # ============ ADMIN — USERS ============
    ("Admin Users", "List", "Admin lists users with filters",
     "GET /api/admin/users?role=Seller&status=Approved",
     "200 paginated", "High"),
    ("Admin Users", "Create", "Admin creates a SalesMan",
     "POST /api/admin/users/salesmen",
     "201; user assigned to specified cities", "High"),
    ("Admin Users", "Create", "Admin creates a ZoneManager",
     "POST /api/admin/users/zonemanagers",
     "201; user assigned to specified region", "High"),
    ("Admin Users", "Toggle Status", "Admin disables a user",
     "PUT /api/admin/users/{id}/toggle-status to disable",
     "200; user.IsDisabled=true; refresh tokens revoked; FcmToken cleared", "High"),
    ("Admin Users", "Toggle Status", "Admin re-enables a user",
     "PUT /api/admin/users/{id}/toggle-status",
     "200; IsDisabled=false (no token side-effects)", "Medium"),
    ("Admin Users", "Delete SM (Reassign)", "Delete SM with pending users — must reassign first",
     "DELETE /api/admin/users/salesmen/{id} without reassignment",
     "400 — must reassign pending users before delete", "High"),
    ("Admin Users", "Delete SM (Reassign)", "Delete SM with reassignment payload",
     "DELETE with newSalesManId for pending users",
     "200; pending users moved to new SM; old SM marked deleted; SM's tokens revoked + FCM cleared", "High"),
    ("Admin Users", "Delete ZM (Reassign)", "Delete ZM — must reassign region's ZM first",
     "DELETE /api/admin/users/zonemanagers/{id} with replacement",
     "200; region.ZoneManagerId reassigned; old ZM tokens revoked", "High"),
    ("Admin Users", "Edit", "Edit SM strips assignments + mobile",
     "PUT /api/admin/users/salesmen/{id}",
     "Per design: Edit clears assignments + mobile (forces explicit reassign step)", "Medium"),

    # ============ ADMIN — PRODUCTS ============
    ("Admin Products", "Create", "Create a product",
     "POST /api/admin/products with valid fields",
     "201", "High"),
    ("Admin Products", "Update", "Update product PointValue",
     "PUT /api/admin/products/{id}",
     "200; existing barcodes still reference original (not retroactive)", "Medium"),
    ("Admin Products", "Delete", "Delete product with no barcodes",
     "DELETE /api/admin/products/{id}",
     "200", "Medium"),
    ("Admin Products", "Delete", "Delete product that has generated barcodes",
     "DELETE on product with existing barcodes",
     "400 — blocked", "High"),

    # ============ ADMIN — BARCODES ============
    ("Admin Barcodes", "Generate", "Generate batch of barcodes (returns PDF)",
     "POST /api/admin/barcodes/generate quantity=100",
     "200 application/pdf with CODE_128 labels (Zebra layout)", "High"),
    ("Admin Barcodes", "Generate", "Generate batch over 1000 limit",
     "POST with quantity=1500",
     "400 — max 1000 per request", "Medium"),
    ("Admin Barcodes", "List", "List barcodes with filters",
     "GET /api/admin/barcodes?productId=...&state=Available",
     "200 paginated", "Medium"),
    ("Admin Scans", "List", "List scan records",
     "GET /api/admin/scans",
     "200 paginated", "Medium"),

    # ============ ADMIN — REWARD SETTINGS ============
    ("Admin Settings", "Get", "Get reward settings",
     "GET /api/admin/reward-settings",
     "200 with InviterRewardPoints, InviteeRewardPoints, MinimumRedemptionPoints, PointsToSarRate", "High"),
    ("Admin Settings", "Update", "Update PointsToSarRate",
     "PUT /api/admin/reward-settings",
     "200; new rate applies to FUTURE earns only (immutability of past txns)", "High"),
    ("Admin Settings", "Update", "Update MinimumRedemptionPoints",
     "PUT settings minimum=2000",
     "Subsequent /api/redemption/cash uses new minimum", "Medium"),

    # ============ ADMIN — ANALYTICS ============
    ("Admin Analytics", "Dashboard", "GET /api/admin/analytics/dashboard",
     "Hit endpoint",
     "200 with KPIs (no 500 — EF Core 10 record-projection bug fixed)", "High"),
    ("Admin Analytics", "Users", "Users analytics",
     "GET /api/admin/analytics/users",
     "200 with breakdown by role / status", "Medium"),
    ("Admin Analytics", "Regions", "Regions analytics",
     "GET /api/admin/analytics/regions",
     "200 with per-region rollup", "Medium"),
    ("Admin Analytics", "Points", "Points summary",
     "GET /api/admin/analytics/points",
     "200 with totals", "Medium"),
    ("Admin Analytics", "Top Performers", "Top performers list",
     "GET /api/admin/analytics/top-performers",
     "200 with top scanners by points", "Medium"),
    ("Admin Analytics", "Inactive Users", "Inactive users",
     "GET /api/admin/analytics/inactive-users",
     "200 with users not active in window", "Low"),
    ("Admin Analytics", "Revenue", "Revenue analytics with date filter",
     "GET /api/admin/analytics/revenue?from=...&to=...",
     "200; date filter respected", "Medium"),
    ("Admin Analytics", "SalesMan Performance", "SM performance",
     "GET /api/admin/analytics/salesman-performance",
     "200 grouped by SM", "Low"),

    # ============ ADMIN — NOTIFICATIONS ============
    ("Admin Notifications", "Send to User", "Admin sends notification to a single user",
     "POST /api/admin/notifications/user with userId + body",
     "200; notification appears in user's list", "High"),
    ("Admin Notifications", "Send to Role", "Admin sends to all Sellers",
     "POST /api/admin/notifications/role role=Seller",
     "200; only active (non-disabled, non-deleted) Sellers receive", "High"),
    ("Admin Notifications", "Send to All", "Admin broadcasts",
     "POST /api/admin/notifications/all",
     "200; all active users; deleted/disabled excluded", "High"),
    ("Admin Notifications", "History", "List notification history",
     "GET /api/admin/notifications?page=1",
     "200 paginated with filters", "Medium"),

    # ============ LOOKUPS ============
    ("Lookups", "Regions", "Public regions list",
     "GET /api/lookup/regions",
     "200 with 8 regions", "Medium"),
    ("Lookups", "Cities", "Public cities filtered by region",
     "GET /api/lookup/cities?regionId=X",
     "200 with cities for region", "Medium"),
    ("Lookups", "Cities", "Hail cities sit under Qassim region",
     "GET /api/lookup/cities?regionId=Qassim",
     "Hail cities present (intentional org structure — see project_hail_under_qassim memory)", "Low"),

    # ============ INFRA / CROSS-CUTTING ============
    ("Infra", "Twilio Mock Mode", "Mock '123456' works in Dev",
     "Login attempt in Dev with Otp=123456",
     "Login succeeds — mock bypass active", "Medium"),
    ("Infra", "Twilio Mock Mode", "Mock NOT allowed in UAT/Production",
     "Set UseMockMode=true in UAT and start app",
     "App fails to start — InvalidOperationException at startup", "High"),
    ("Infra", "OTP Lifetime", "OTP valid for 10 minutes (was 3)",
     "Send OTP, wait 9 minutes, attempt verify",
     "200 — within window", "Medium"),
    ("Infra", "OTP Lifetime", "OTP expired after 10 minutes",
     "Wait >10 min then verify",
     "400 — OTP expired", "Medium"),
    ("Infra", "Verification Token", "Verification token valid for 5 hours",
     "Hold the post-OTP token for ~4h then complete registration",
     "Registration completes within 5h window", "Low"),
    ("Infra", "WhatsApp Send", "Cash OTP template uses 3 vars + whatsapp: prefix on From",
     "Trigger cash OTP, watch Twilio logs",
     "Send succeeds with Content {1}=otp {2}=points {3}=sar; both From and To prefixed whatsapp:", "High"),
    ("Infra", "Auto-Migration", "App auto-migrates on Dev/Staging/UAT startup",
     "Bring up env with newer migration than DB",
     "Migration applies on startup; no manual step", "Medium"),
    ("Infra", "JWT Validation", "Missing JWT key fails app startup",
     "Remove Jwt:Key from appsettings, start app",
     "App refuses to start with config validation error", "High"),
]

# ----- workbook -----
wb = Workbook()

# === Test Cases sheet ===
ws = wb.active
ws.title = "Test Cases"
headers = [
    "TC-ID", "Feature", "Sub-Feature", "Scenario", "Steps",
    "Expected Result", "Priority", "Status", "Tester", "Bug ID", "Notes"
]
ws.append(headers)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER

for i, (feat, sub, scen, steps, expected, prio) in enumerate(CASES, start=1):
    tc_id = f"TC-{i:03d}"
    row = [tc_id, feat, sub, scen, steps, expected, prio, "Not Started", "", "", ""]
    ws.append(row)
    r = i + 1
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        cell.border = BORDER
    # priority colour
    p_cell = ws.cell(row=r, column=7)
    p_cell.fill = PRIO_FILLS[prio]
    p_cell.alignment = CENTER

# column widths
widths = {
    "A": 9, "B": 20, "C": 22, "D": 45, "E": 55,
    "F": 55, "G": 11, "H": 14, "I": 14, "J": 12, "K": 30,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# data validation: Status
status_dv = DataValidation(
    type="list",
    formula1='"Not Started,Pass,Fail,Blocked,N/A"',
    allow_blank=True,
)
status_dv.add(f"H2:H{len(CASES) + 1}")
ws.add_data_validation(status_dv)

# data validation: Priority
prio_dv = DataValidation(
    type="list",
    formula1='"High,Medium,Low"',
    allow_blank=False,
)
prio_dv.add(f"G2:G{len(CASES) + 1}")
ws.add_data_validation(prio_dv)

# conditional formatting on Status column (H)
status_range = f"H2:H{len(CASES) + 1}"
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Pass"'],
    fill=PatternFill("solid", fgColor="C6E0B4")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Fail"'],
    fill=PatternFill("solid", fgColor="F4B084")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Blocked"'],
    fill=PatternFill("solid", fgColor="FFE699")))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"N/A"'],
    fill=PatternFill("solid", fgColor="D9D9D9")))

# auto-filter + freeze
ws.auto_filter.ref = f"A1:K{len(CASES) + 1}"
ws.freeze_panes = "B2"

# === Summary sheet ===
summary = wb.create_sheet("Summary", 0)
summary["A1"] = "RewardProgram — QA Test Plan"
summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
summary["A2"] = f"Total test cases: {len(CASES)}"
summary["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")

# feature roll-up
features = {}
for feat, _, _, _, _, prio in CASES:
    features.setdefault(feat, {"High": 0, "Medium": 0, "Low": 0, "Total": 0})
    features[feat][prio] += 1
    features[feat]["Total"] += 1

hdr = ["Feature", "High", "Medium", "Low", "Total"]
for col, h in enumerate(hdr, start=1):
    cell = summary.cell(row=4, column=col, value=h)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER

row = 5
for feat in sorted(features.keys()):
    counts = features[feat]
    summary.cell(row=row, column=1, value=feat).font = BODY_FONT
    summary.cell(row=row, column=2, value=counts["High"]).fill = PRIO_FILLS["High"]
    summary.cell(row=row, column=3, value=counts["Medium"]).fill = PRIO_FILLS["Medium"]
    summary.cell(row=row, column=4, value=counts["Low"]).fill = PRIO_FILLS["Low"]
    summary.cell(row=row, column=5, value=counts["Total"]).font = Font(bold=True)
    for c in range(1, 6):
        cc = summary.cell(row=row, column=c)
        cc.alignment = CENTER if c > 1 else WRAP
        cc.border = BORDER
    row += 1

# totals
summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
summary.cell(row=row, column=2, value=sum(f["High"] for f in features.values())).font = Font(bold=True)
summary.cell(row=row, column=3, value=sum(f["Medium"] for f in features.values())).font = Font(bold=True)
summary.cell(row=row, column=4, value=sum(f["Low"] for f in features.values())).font = Font(bold=True)
summary.cell(row=row, column=5, value=len(CASES)).font = Font(bold=True)
for c in range(1, 6):
    summary.cell(row=row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
    summary.cell(row=row, column=c).border = BORDER
    summary.cell(row=row, column=c).alignment = CENTER if c > 1 else WRAP

summary.column_dimensions["A"].width = 26
for col in ("B", "C", "D", "E"):
    summary.column_dimensions[col].width = 12

# === Legend sheet ===
legend = wb.create_sheet("Legend")
legend["A1"] = "Legend"
legend["A1"].font = Font(size=14, bold=True, color="1F4E78")

legend["A3"] = "Priority"
legend["A3"].font = Font(bold=True)
legend["A4"] = "High"
legend["A4"].fill = PRIO_FILLS["High"]
legend["B4"] = "Critical path / data-integrity / security — must pass before release"
legend["A5"] = "Medium"
legend["A5"].fill = PRIO_FILLS["Medium"]
legend["B5"] = "Important but not blocking — should pass before release"
legend["A6"] = "Low"
legend["A6"].fill = PRIO_FILLS["Low"]
legend["B6"] = "Edge case / nice-to-have — log defects but ship-allowed"

legend["A8"] = "Status"
legend["A8"].font = Font(bold=True)
legend["A9"] = "Not Started"
legend["B9"] = "Default — test not yet executed"
legend["A10"] = "Pass"
legend["A10"].fill = PatternFill("solid", fgColor="C6E0B4")
legend["B10"] = "Test executed and matches expected result"
legend["A11"] = "Fail"
legend["A11"].fill = PatternFill("solid", fgColor="F4B084")
legend["B11"] = "Test executed and result does NOT match expected — log Bug ID"
legend["A12"] = "Blocked"
legend["A12"].fill = PatternFill("solid", fgColor="FFE699")
legend["B12"] = "Cannot execute — dependency / env issue. Note in Notes column"
legend["A13"] = "N/A"
legend["A13"].fill = PatternFill("solid", fgColor="D9D9D9")
legend["B13"] = "Not applicable in this environment / build"

legend.column_dimensions["A"].width = 14
legend.column_dimensions["B"].width = 80

# save
out = Path(__file__).parent.parent / "docs" / "QA-TestPlan.xlsx"
wb.save(out)
print(f"Wrote {out} with {len(CASES)} test cases across {len(features)} features.")
